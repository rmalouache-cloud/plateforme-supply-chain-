import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side
from io import BytesIO
from PIL import Image
import matplotlib.pyplot as plt
from reportlab.platypus import SimpleDocTemplate, Image as RLImage
import tempfile
import time

# ==============================
# CONFIG
# ==============================
st.set_page_config(page_title="BOM Comparator", layout="wide")

# ==============================
# LOGO
# ==============================
try:
    logo = Image.open("logo.jfif")
    st.image(logo, width=1500)
except:
    st.title("BOM Comparator")

st.markdown("## 📊  BOM vs Packing Comparison Tool  ⚖️")

# ==============================
# INPUTS
# ==============================
bom_file = st.file_uploader("📄  Upload BOM file", type=["xlsx", "xls"])
packing_file = st.file_uploader("📦 Upload Packing file", type=["xlsx", "xls"])

model_input = st.text_input("📺 Enter Model")
lot_input = st.text_input(" 🔢 Enter Lot Quantity")

run = st.button("🚀 Compare")

# ==============================
# KPI
# ==============================
def show_kpis(df):
    total = len(df)
    
    # Compter les articles uniques avec changement de référence
    # Chaque paire de changement compte comme 1 dans le compteur
    ref_change_pairs = len(st.session_state.get("ref_changes", {}))
    
    conform = (df["Remark"] == "✅ Conform").sum()
    missing = (df["Remark"] == "❌ Missing item").sum()
    packing_only = (df["Remark"] == "📦 Packing only").sum()
    qty_missing = (df["Remark"] == "⚠ Qty missing").sum()
    
    # Afficher le nombre de paires de changement, pas le nombre d'articles
    st.markdown(f"### 📊 Total Articles: {total}")

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("✅ Conform", conform)
    c2.metric("❌ Missing", missing)
    c3.metric("📦 Packing only", packing_only)
    c4.metric("⚠ Qty missing", qty_missing)
    c5.metric("🔄 Ref Change", ref_change_pairs)

# ==============================
# PIE CHART
# ==============================
def generate_pie_chart(df):
    conform = (df["Remark"] == "✅ Conform").sum()
    missing = (df["Remark"] == "❌ Missing item").sum()
    packing_only = (df["Remark"] == "📦 Packing only").sum()
    qty_missing = (df["Remark"] == "⚠ Qty missing").sum()
    ref_change_pairs = len(st.session_state.get("ref_changes", {}))
    
    # Pour le graphique, on utilise le nombre de paires
    labels = ["Conform", "Missing", "Packing Only", "Qty Missing", "Ref Change"]
    values = [conform, missing, packing_only, qty_missing, ref_change_pairs]
    colors = ['#4CAF50', '#F44336', '#9C27B0', '#FF9800', '#2196F3']

    fig, ax = plt.subplots(figsize=(4, 4))
    ax.pie(values, labels=labels, autopct="%1.1f%%", startangle=90, colors=colors)
    ax.set_title("KPI Distribution (Articles)")
    return fig

# ==============================
# TABLE STYLE avec bordures noires
# ==============================
def highlight_remark_column(df):
    styles = []
    for val in df["Remark"]:
        if val == "✅ Conform":
            styles.append("background-color: #4CAF50; color: white; font-weight: bold; border: 2px solid #000000; border-radius: 5px;")
        elif val == "⚠ Qty missing":
            styles.append("background-color: #FF9800; color: white; font-weight: bold; border: 2px solid #000000; border-radius: 5px;")
        elif val == "❌ Missing item":
            styles.append("background-color: #F44336; color: white; font-weight: bold; border: 2px solid #000000; border-radius: 5px;")
        elif val == "📦 Packing only":
            styles.append("background-color: #9C27B0; color: white; font-weight: bold; border: 2px solid #000000; border-radius: 5px;")
        elif val == "🔄 Reference Change":
            styles.append("background-color: #2196F3; color: white; font-weight: bold; border: 2px solid #000000; border-radius: 5px;")
        else:
            styles.append("")
    
    style_df = pd.DataFrame("", index=df.index, columns=df.columns)
    style_df["Remark"] = styles
    return style_df

# ==============================
# EXCEL EXPORT AVEC COULEURS CLAIRES ET BORDURES NOIRES
# ==============================
def export_excel(df, common_equipment_df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Résultats détaillés")
        if not common_equipment_df.empty:
            common_equipment_df.to_excel(writer, index=False, sheet_name="Équipements communs")
    output.seek(0)
    wb = load_workbook(output)
    
    # Définir les bordures noires
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Styliser la première feuille (Résultats détaillés)
    ws1 = wb["Résultats détaillés"]
    for row in ws1.iter_rows():
        for cell in row:
            cell.border = thin_border
    
    # Couleurs plus claires pour Excel
    color_map = {
        "✅ Conform": "C6EFCE",      # Vert très clair
        "⚠ Qty missing": "FFEB9C",   # Orange très clair
        "❌ Missing item": "FFC7CE",  # Rouge très clair
        "📦 Packing only": "E6D0FF",  # Violet très clair
        "🔄 Reference Change": "B3D9FF"  # Bleu très clair
    }
    
    # Appliquer les couleurs aux lignes
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
        remark_cell = row[8]  # Colonne Remark (index 8)
        if remark_cell.value in color_map:
            color = color_map[remark_cell.value]
            for cell in row:
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    
    # Ajuster la largeur des colonnes pour la première feuille
    for column in ws1.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws1.column_dimensions[column_letter].width = adjusted_width
    
    # Styliser la deuxième feuille (Équipements communs) si elle existe
    if not common_equipment_df.empty and "Équipements communs" in wb.sheetnames:
        ws2 = wb["Équipements communs"]
        for row in ws2.iter_rows():
            for cell in row:
                cell.border = thin_border
        
        # Ajuster la largeur des colonnes pour la deuxième feuille
        for column in ws2.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws2.column_dimensions[column_letter].width = adjusted_width
    
    final = BytesIO()
    wb.save(final)
    final.seek(0)
    return final

# ==============================
# MAIN CALCULATION
# ==============================
if run:
    if not bom_file or not packing_file:
        st.error("Upload both files")
        st.stop()
    if not model_input:
        st.error("Enter model")
        st.stop()
    if not lot_input.isdigit():
        st.error("Lot must be numeric")
        st.stop()
    lot = int(lot_input)
    bom = pd.read_excel(bom_file)
    packing = pd.read_excel(packing_file)
    bom.columns = bom.columns.str.strip()
    packing.columns = packing.columns.str.strip()
    
    # Nettoyer la colonne Model
    packing["Model"] = packing["Model"].astype(str).str.strip()
    packing["Model"] = packing["Model"].replace("", None).replace("-", None).replace("nan", None)
    
    # Séparer les équipements communs (sans modèle) des équipements spécifiques
    common_equipment = packing[packing["Model"].isna()].copy()
    packing_with_model = packing[packing["Model"].notna()].copy()
    
    # Propager les modèles pour les lignes vides (méthode ffill)
    packing_with_model["Model"] = packing_with_model["Model"].ffill()
    
    # Filtrer par modèle sélectionné
    packing_model = packing_with_model[packing_with_model["Model"] == model_input]
    
    if packing_model.empty and common_equipment.empty:
        st.error("Model not found and no common equipment")
        st.stop()
    
    # Traitement des équipements communs
    common_equipment_summary = pd.DataFrame()
    if not common_equipment.empty:
        common_equipment_summary = common_equipment.groupby(["PN", "Description"])["packing_qty"].sum().reset_index()
        common_equipment_summary = common_equipment_summary.rename(columns={
            "packing_qty": "Qty expédiée"
        })
    
    # Traitement des BOM vs Packing pour le modèle sélectionné
    if not packing_model.empty:
        bom_g = bom.groupby(["PN", "Description"])["bom_qty"].sum().reset_index()
        packing_g = packing_model.groupby(["PN", "Description"])["packing_qty"].sum().reset_index()
        df = pd.merge(bom_g, packing_g, on="PN", how="outer", suffixes=("_BOM", "_Packing"), indicator=True)
        df["bom_qty"] = pd.to_numeric(df["bom_qty"], errors="coerce").fillna(0)
        df["packing_qty"] = pd.to_numeric(df["packing_qty"], errors="coerce").fillna(0)
        df["Description_BOM"] = df["Description_BOM"].fillna(df["Description_Packing"])
        df["MP"] = df["bom_qty"] * lot
        df["SAV"] = df["MP"] * 0.02
        df["Qty (MP+SAV)"] = df["MP"] + df["SAV"]
        df["Balance"] = df["packing_qty"] - df["Qty (MP+SAV)"]
        
        def detect_remark(row):
            if row["_merge"] == "left_only":
                return "❌ Missing item"
            elif row["_merge"] == "right_only":
                return "📦 Packing only"
            elif row["packing_qty"] >= row["Qty (MP+SAV)"]:
                return "✅ Conform"
            else:
                return "⚠ Qty missing"
        
        df["Remark"] = df.apply(detect_remark, axis=1)
        result = df[[
            "PN", "Description_BOM", "bom_qty", "packing_qty",
            "MP", "SAV", "Qty (MP+SAV)", "Balance", "Remark"
        ]].rename(columns={
            "Description_BOM": "Description",
            "bom_qty": "Qty BOM",
            "packing_qty": "Packing list qty"
        })
    else:
        result = pd.DataFrame()  # DataFrame vide si pas de données pour le modèle
    
    st.session_state["result"] = result
    st.session_state["common_equipment"] = common_equipment_summary
    st.session_state["data_ready"] = True
    # Initialiser les changements de référence
    if "ref_changes" not in st.session_state:
        st.session_state["ref_changes"] = {}

# ==============================
# AFFICHAGE DES RÉSULTATS
# ==============================
if "data_ready" in st.session_state and st.session_state["data_ready"]:
    result = st.session_state["result"].copy() if not st.session_state["result"].empty else pd.DataFrame()
    common_equipment = st.session_state["common_equipment"]
    
    # Appliquer les changements de référence
    if not result.empty:
        for old_ref, new_ref in st.session_state["ref_changes"].items():
            result.loc[result["PN"] == old_ref, "Remark"] = "🔄 Reference Change"
            result.loc[result["PN"] == new_ref, "Remark"] = "🔄 Reference Change"
    
    st.success("Comparison completed ✅")
    
    # 1. KPIs (seulement si result non vide)
    if not result.empty:
        show_kpis(result)
        st.markdown("---")
    
    # 2. TABLEAU DES ÉQUIPEMENTS COMMUNS (avant le tableau principal)
    if not common_equipment.empty:
        st.markdown("### 🛠️ Équipements communs expédiés avec la marchandise")
        st.markdown("*Ces équipements sont livrés avec tous les modèles (pas de modèle spécifique)*")
        
        styled_common = common_equipment.style.set_properties(**{
            'background-color': '#E3F2FD',
            'color': '#0D47A1',
            'font-weight': 'bold',
            'border': '2px solid #000000',
            'border-radius': '5px'
        })
        st.dataframe(styled_common, use_container_width=True)
        st.markdown("---")
    
    # 3. TABLEAU PRINCIPAL (si non vide)
    if not result.empty:
        styled = result.style.apply(highlight_remark_column, axis=None)
        st.dataframe(styled, use_container_width=True)
        st.markdown("---")
    else:
        st.info("ℹ️ Aucune donnée trouvée pour le modèle sélectionné")
        st.markdown("---")
    
    # 4. GESTION CHANGEMENT REFERENCE (seulement si result non vide)
    if not result.empty:
        st.markdown("### 🔄 Gestion des changements de référence")
        
        # Créer trois onglets pour mieux organiser
        tab1, tab2, tab3 = st.tabs(["➕ Nouveau changement", "📋 Changements actifs", "ℹ️ Aide"])
        
        with tab1:
            missing_items = result[result["Remark"] == "❌ Missing item"]
            packing_items = result[result["Remark"] == "📦 Packing only"]
            
            if missing_items.empty or packing_items.empty:
                st.warning("⚠️ Aucun article 'Missing' ou 'Packing only' détecté pour créer un changement de référence")
            else:
                # Ajouter des statistiques
                col_stat1, col_stat2 = st.columns(2)
                with col_stat1:
                    st.info(f"❌ **{len(missing_items)}** article(s) 'Missing' disponibles")
                with col_stat2:
                    st.info(f"📦 **{len(packing_items)}** article(s) 'Packing only' disponibles")
                
                st.markdown("---")
                
                # Sélection avec recherche et filtrage
                col1, col2 = st.columns(2)
                
                with col1:
                    st.markdown("**❌ Ancienne référence (Missing)**")
                    
                    # Ajouter une barre de recherche
                    search_missing = st.text_input("🔍 Rechercher", placeholder="PN ou Description...", key="search_missing")
                    
                    # Filtrer les résultats
                    missing_filtered = missing_items
                    if search_missing:
                        missing_filtered = missing_items[
                            missing_items["PN"].str.contains(search_missing, case=False, na=False) |
                            missing_items["Description"].str.contains(search_missing, case=False, na=False)
                        ]
                    
                    if not missing_filtered.empty:
                        selected_missing = st.selectbox(
                            "Sélectionner l'article manquant",
                            options=missing_filtered["PN"].tolist(),
                            format_func=lambda x: f"🔴 {x} - {missing_filtered[missing_filtered['PN']==x]['Description'].values[0][:60]}",
                            key="missing_select"
                        )
                        
                        # Afficher les détails de l'article sélectionné
                        if selected_missing:
                            missing_details = missing_filtered[missing_filtered["PN"] == selected_missing].iloc[0]
                            st.caption(f"📝 {missing_details['Description'][:100]}")
                            st.caption(f"📊 Quantité BOM: {missing_details['Qty BOM']:.0f}")
                    else:
                        st.warning("Aucun résultat trouvé")
                        selected_missing = None
                
                with col2:
                    st.markdown("**📦 Nouvelle référence (Packing only)**")
                    
                    # Ajouter une barre de recherche
                    search_packing = st.text_input("🔍 Rechercher", placeholder="PN ou Description...", key="search_packing")
                    
                    # Filtrer les résultats
                    packing_filtered = packing_items
                    if search_packing:
                        packing_filtered = packing_items[
                            packing_items["PN"].str.contains(search_packing, case=False, na=False) |
                            packing_items["Description"].str.contains(search_packing, case=False, na=False)
                        ]
                    
                    if not packing_filtered.empty:
                        selected_packing = st.selectbox(
                            "Sélectionner le nouvel article",
                            options=packing_filtered["PN"].tolist(),
                            format_func=lambda x: f"🔵 {x} - {packing_filtered[packing_filtered['PN']==x]['Description'].values[0][:60]}",
                            key="packing_select"
                        )
                        
                        # Afficher les détails de l'article sélectionné
                        if selected_packing:
                            packing_details = packing_filtered[packing_filtered["PN"] == selected_packing].iloc[0]
                            st.caption(f"📝 {packing_details['Description'][:100]}")
                            st.caption(f"📊 Quantité Packing: {packing_details['Packing list qty']:.0f}")
                    else:
                        st.warning("Aucun résultat trouvé")
                        selected_packing = None
                
                # Bouton d'application avec vérification
                if selected_missing and selected_packing:
                    st.markdown("---")
                    
                    # Vérifier si le changement existe déjà
                    already_exists = False
                    existing_pair = None
                    for old, new in st.session_state["ref_changes"].items():
                        if old == selected_missing:
                            already_exists = True
                            existing_pair = (old, new)
                            break
                        if new == selected_packing:
                            already_exists = True
                            existing_pair = (old, new)
                            break
                    
                    if already_exists:
                        st.warning(f"⚠️ Cet article est déjà impliqué dans un changement : {existing_pair[0]} → {existing_pair[1]}")
                    else:
                        # Aperçu du changement
                        st.info(f"**Aperçu :** {selected_missing} → {selected_packing}")
                        
                        col_btn1, col_btn2, col_btn3 = st.columns([1, 2, 1])
                        with col_btn2:
                            if st.button("✅ Appliquer ce changement", use_container_width=True, type="primary"):
                                st.session_state["ref_changes"][selected_missing] = selected_packing
                                st.success(f"✅ Changement appliqué avec succès ! {selected_missing} → {selected_packing}")
                                time.sleep(0.5)
                                st.rerun()
        
        with tab2:
            if st.session_state["ref_changes"]:
                st.markdown("### 📝 Changements de référence effectués")
                
                # Créer un dataframe pour l'affichage
                changes_list = []
                for idx, (old, new) in enumerate(st.session_state["ref_changes"].items(), 1):
                    # Récupérer les descriptions
                    old_desc = result[result["PN"] == old]["Description"].values[0] if old in result["PN"].values else "N/A"
                    new_desc = result[result["PN"] == new]["Description"].values[0] if new in result["PN"].values else "N/A"
                    
                    changes_list.append({
                        "#": idx,
                        "Ancienne référence": old,
                        "Ancienne description": old_desc[:50],
                        "Nouvelle référence": new,
                        "Nouvelle description": new_desc[:50]
                    })
                
                changes_df = pd.DataFrame(changes_list)
                st.dataframe(changes_df, use_container_width=True, hide_index=True)
                
                # Statistiques des changements
                st.markdown("---")
                col_stat1, col_stat2, col_stat3 = st.columns(3)
                with col_stat1:
                    st.metric("Total changements", len(st.session_state["ref_changes"]))
                with col_stat2:
                    st.metric("Articles impactés", len(st.session_state["ref_changes"]) * 2)
                with col_stat3:
                    total_problems = len(result[result['Remark'].isin(['❌ Missing item', '📦 Packing only'])])
                    if total_problems > 0:
                        correction_rate = (len(st.session_state["ref_changes"]) / total_problems) * 100
                        st.metric("Taux de correction", f"{correction_rate:.1f}%")
                    else:
                        st.metric("Taux de correction", "0%")
                
                # Bouton de réinitialisation
                st.markdown("---")
                col_reset1, col_reset2, col_reset3 = st.columns([1, 2, 1])
                with col_reset2:
                    if st.button("🗑️ Réinitialiser tous les changements", use_container_width=True):
                        st.session_state["ref_changes"] = {}
                        st.success("✅ Tous les changements ont été réinitialisés")
                        time.sleep(0.5)
                        st.rerun()
            else:
                st.info("💡 Aucun changement de référence n'a encore été appliqué")
                st.markdown("""
                **Comment créer un changement ?**
                
                1. Allez dans l'onglet '➕ Nouveau changement'
                2. Sélectionnez un article 'Missing' (ancienne référence)
                3. Sélectionnez un article 'Packing only' (nouvelle référence)
                4. Cliquez sur 'Appliquer ce changement'
                """)
        
        with tab3:
            st.markdown("### ℹ️ Guide d'utilisation")
            
            col_help1, col_help2 = st.columns(2)
            
            with col_help1:
                st.markdown("**🎯 Quand utiliser cette fonction ?**")
                st.markdown("- Lorsqu'une référence a été changée/remplacée dans la nouvelle version")
                st.markdown("- Quand un article 'Missing' et un article 'Packing only' correspondent au même composant")
                st.markdown("- Pour fusionner deux lignes qui représentent le même article")
                st.markdown("")
                st.markdown("**📊 Impact sur les KPIs**")
                st.markdown("- Chaque changement de référence compte comme +1 dans 'Ref Change'")
                st.markdown("- Les articles deviennent '🔄 Reference Change' dans le tableau")
                st.markdown("- Le graphique circulaire est automatiquement mis à jour")
            
            with col_help2:
                st.markdown("**💡 Conseils**")
                st.markdown("- Utilisez la barre de recherche pour trouver rapidement un article")
                st.markdown("- Vérifiez les descriptions pour confirmer la correspondance")
                st.markdown("- Un article ne peut pas être utilisé dans deux changements différents")
                st.markdown("- Vous pouvez annuler tous les changements à tout moment")
                st.markdown("")
                st.markdown("**🔍 Exemple**")
                st.markdown("```")
                st.markdown("Ancienne réf: RES-100 (Missing)")
                st.markdown("Nouvelle réf: RES-200 (Packing only)")
                st.markdown("→ Même résistance, référence changée")
                st.markdown("```")
            
            st.markdown("---")
            st.info("💬 Astuce : Les changements de référence ne modifient pas les quantités, ils permettent juste de suivre l'évolution des références.")
        
        st.markdown("---")
    
    # 5. CERCLE APRÈS TABLEAU (seulement si result non vide)
    if not result.empty:
        st.markdown("### 📊 KPI Distribution")
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            fig = generate_pie_chart(result)
            st.pyplot(fig)
            
            # PDF Export
            img_buffer = BytesIO()
            fig.savefig(img_buffer, format="png")
            img_buffer.seek(0)
            pdf_buffer = BytesIO()
            doc = SimpleDocTemplate(pdf_buffer)
            with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp:
                tmp.write(img_buffer.getvalue())
                tmp_path = tmp.name
            elements = [RLImage(tmp_path, width=300, height=300)]
            doc.build(elements)
            pdf_buffer.seek(0)
            st.download_button(
                "📄 Download KPI Chart (PDF)",
                data=pdf_buffer,
                file_name="KPI_Chart.pdf",
                mime="application/pdf"
            )
    
    # EXCEL DOWNLOAD
    excel_file = export_excel(result if not result.empty else pd.DataFrame(), common_equipment)
    st.download_button(
        "📥 Download Excel Result",
        data=excel_file,
        file_name="BOM_vs_Packing.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
