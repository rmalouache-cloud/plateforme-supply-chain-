import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment

# Configuration
st.set_page_config(
    page_title="Vérification Fournisseur",
    page_icon="✅",
    layout="wide"
)

# ==================== CSS ====================
st.markdown("""
<style>
    .header {
        background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%);
        padding: 2rem;
        border-radius: 15px;
        margin-bottom: 2rem;
        text-align: center;
    }
    
    .header h1 {
        color: white;
        margin: 0;
        font-size: 2rem;
    }
    
    .header p {
        color: #e0e0e0;
        margin: 0.5rem 0 0 0;
    }
    
    .card {
        background: white;
        border-radius: 12px;
        padding: 1.5rem;
        margin-bottom: 1.5rem;
        border: 1px solid #e0e0e0;
        box-shadow: 0 2px 4px rgba(0,0,0,0.05);
    }
    
    .card-title {
        font-size: 1.2rem;
        font-weight: 600;
        margin-bottom: 1rem;
        color: #1e3c72;
        display: flex;
        align-items: center;
        gap: 0.5rem;
    }
    
    .upload-area {
        border: 2px dashed #2a5298;
        border-radius: 10px;
        padding: 1.5rem;
        text-align: center;
        background: #fafafa;
    }
    
    .metric-total { background: #3b82f6; border-radius: 10px; padding: 1rem; text-align: center; color: white; }
    .metric-correct { background: #10b981; border-radius: 10px; padding: 1rem; text-align: center; color: white; }
    .metric-incorrect { background: #ef4444; border-radius: 10px; padding: 1rem; text-align: center; color: white; }
    .metric-taux { background: #f59e0b; border-radius: 10px; padding: 1rem; text-align: center; color: white; }
    .metric-number { font-size: 2rem; font-weight: bold; }
    .metric-label { font-size: 0.8rem; margin-top: 0.3rem; }
    
    .info-box {
        background: #e0e7ff;
        padding: 0.8rem;
        border-radius: 8px;
        margin-bottom: 1rem;
        color: #1e3c72;
        font-weight: 500;
    }
    
    .footer {
        text-align: center;
        padding: 1.5rem;
        margin-top: 2rem;
        border-top: 1px solid #e0e0e0;
        color: #666;
    }
    
    .stButton > button {
        background: linear-gradient(135deg, #1e3c72 0%, #2a5298 100%);
        color: white;
        border: none;
        padding: 0.7rem 3rem;
        font-weight: 600;
        border-radius: 8px;
        font-size: 1rem;
        min-width: 250px;
    }
    
    .stTextInput > div > div > input {
        border-radius: 8px;
        border: 1px solid #d0d0d0;
        padding: 0.5rem;
    }
    
    .result-table {
        width: 100%;
        border-collapse: collapse;
        font-size: 12px;
        overflow-x: auto;
        display: block;
    }
    
    .result-table th {
        background-color: #1e3c72;
        color: white;
        padding: 10px 8px;
        border: 1px solid #2a5298;
        font-weight: 600;
        white-space: nowrap;
    }
    
    .result-table td {
        border: 1px solid #ddd;
        padding: 8px;
        white-space: nowrap;
    }
    
    .result-table tr:hover {
        background-color: #f5f5f5;
    }
    
    .status-success {
        background-color: #d1fae5;
        color: #065f46;
        font-weight: bold;
        padding: 4px 8px;
        border-radius: 4px;
        display: inline-block;
    }
    
    .status-error {
        background-color: #fee2e2;
        color: #991b1b;
        font-weight: bold;
        padding: 4px 8px;
        border-radius: 4px;
        display: inline-block;
    }
    
    .status-warning {
        background-color: #fed7aa;
        color: #92400e;
        font-weight: bold;
        padding: 4px 8px;
        border-radius: 4px;
        display: inline-block;
    }
</style>
""", unsafe_allow_html=True)

# ==================== EN-TÊTE ====================
st.markdown("""
<div class="header">
    <h1>✅ Vérification des Réponses Fournisseur</h1>
    <p>Contrôle automatique des quantités Oversent vs Stock</p>
</div>
""", unsafe_allow_html=True)

# ==================== FONCTIONS ====================

def charger_feuilles_reply(uploaded_file):
    try:
        xlsx = pd.ExcelFile(uploaded_file)
        return {sheet: pd.read_excel(uploaded_file, sheet_name=sheet) for sheet in xlsx.sheet_names}
    except Exception as e:
        st.error(f"❌ Erreur: {e}")
        return None

def charger_stocks(uploaded_files):
    stocks = {}
    for f in uploaded_files:
        try:
            stocks[f.name] = pd.read_excel(f)
        except Exception as e:
            st.error(f"❌ Erreur {f.name}: {e}")
    return stocks

def extraire_colonnes_reply(df):
    if len(df.columns) < 9:
        return None
    df = df.copy()
    df['Part_N'] = df.iloc[:, 0].astype(str).str.strip()
    df['Description'] = df.iloc[:, 1].astype(str).str.strip()
    df['Packing_qty'] = pd.to_numeric(df.iloc[:, 3], errors='coerce').fillna(0)
    df['Qty_for'] = pd.to_numeric(df.iloc[:, 4], errors='coerce').fillna(0)
    df['Remarks'] = df.iloc[:, 6].astype(str).str.strip()
    df['Moka_file'] = df.iloc[:, 7].astype(str).str.strip().replace('.xlsx', '', regex=True).replace('.xls', '', regex=True)
    df['Oversent_FRS'] = pd.to_numeric(df.iloc[:, 8], errors='coerce').fillna(0)
    return df

def get_oversent_stock(df_stock, part_n, idl):
    if len(df_stock.columns) < 11:
        raise ValueError("Colonnes insuffisantes")
    
    col_part = df_stock.columns[3]
    col_idl = df_stock.columns[0]
    
    mask = df_stock[col_part].astype(str).str.strip() == str(part_n).strip()
    df_filtered = df_stock[mask].reset_index(drop=True)
    
    if df_filtered.empty:
        raise ValueError(f"Part N non trouvé")
    
    mask_idl = df_filtered[col_idl].astype(str).str.strip() == str(idl).strip()
    idx = df_filtered[mask_idl].index
    
    if len(idx) == 0:
        raise ValueError(f"IDL non trouvé")
    
    pos = idx[0]
    if pos == 0:
        raise ValueError(f"IDL à la première ligne")
    
    val = df_filtered.iloc[pos - 1, 10]
    return float(val) if pd.notna(val) else 0.0

def exporter_excel_stylise(df, erreurs):
    """Exporte le DataFrame en Excel avec bordures et couleurs"""
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Résultats', index=False)
        
        workbook = writer.book
        worksheet = writer.sheets['Résultats']
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        header_fill = PatternFill(start_color='1e3c72', end_color='1e3c72', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        success_fill = PatternFill(start_color='d1fae5', end_color='d1fae5', fill_type='solid')
        success_font = Font(color='065f46', bold=True)
        
        error_fill = PatternFill(start_color='fee2e2', end_color='fee2e2', fill_type='solid')
        error_font = Font(color='991b1b', bold=True)
        
        warning_fill = PatternFill(start_color='fed7aa', end_color='fed7aa', fill_type='solid')
        warning_font = Font(color='92400e', bold=True)
        
        for col_idx, column in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        for row_idx in range(2, len(df) + 2):
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center' if isinstance(cell.value, (int, float)) else 'left')
                
                if df.columns[col_idx - 1] == 'Status':
                    if cell.value == '✅':
                        cell.fill = success_fill
                        cell.font = success_font
                    elif cell.value == '❌':
                        cell.fill = error_fill
                        cell.font = error_font
                    elif cell.value == '⚠️':
                        cell.fill = warning_fill
                        cell.font = warning_font
        
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        if erreurs:
            df_erreurs = pd.DataFrame({'Erreurs': erreurs})
            df_erreurs.to_excel(writer, sheet_name='Erreurs', index=False)
            worksheet_erreurs = writer.sheets['Erreurs']
            for row_idx in range(1, len(df_erreurs) + 2):
                for col_idx in range(1, len(df_erreurs.columns) + 1):
                    cell = worksheet_erreurs.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
    
    return output.getvalue()

def afficher_tableau_html(df):
    """Affiche un tableau HTML propre avec bordures et nouveaux noms de colonnes"""
    
    # Renommer les colonnes
    df_aff = df.copy()
    colonnes_rename = {
        'Modèle': 'Modèle',
        'Part N': 'Part N',
        'Description': 'Description',
        'Remarks': 'Remarks',
        'IDL': 'IDL',
        'Qty for': 'Qty need in this lot',
        'Packing Qty': 'Qty sent in this lot',
        'Oversent Stock': 'Oversent in the previous IDL',
        'Oversent FRS': 'Oversent of reply',
        'Oversent Calculé': 'Oversent calculated',
        'Écart': 'GAP',
        'Status': 'Status'
    }
    
    df_aff = df_aff.rename(columns=colonnes_rename)
    
    html = '<div style="overflow-x: auto;">'
    html += '<table class="result-table" style="width: 100%; border-collapse: collapse;">'
    
    # En-têtes
    html += '<thead>'
    html += ' hilab'
    for col in df_aff.columns:
        html += f'<th style="background-color: #1e3c72; color: white; padding: 10px; border: 1px solid #2a5298; font-weight: bold;">{col}</th>'
    html += '</tr>'
    html += '</thead>'
    
    # Corps
    html += '<tbody>'
    for _, row in df_aff.iterrows():
        html += ' hilab'
        for col in df_aff.columns:
            value = row[col]
            
            # Style pour la colonne Status
            if col == 'Status':
                if value == '✅':
                    html += f'<td style="border: 1px solid #ddd; padding: 8px; text-align: center;"><span style="background-color: #d1fae5; color: #065f46; font-weight: bold; padding: 4px 8px; border-radius: 4px;">{value}</span></td>'
                elif value == '❌':
                    html += f'<td style="border: 1px solid #ddd; padding: 8px; text-align: center;"><span style="background-color: #fee2e2; color: #991b1b; font-weight: bold; padding: 4px 8px; border-radius: 4px;">{value}</span></td>'
                elif value == '⚠️':
                    html += f'<td style="border: 1px solid #ddd; padding: 8px; text-align: center;"><span style="background-color: #fed7aa; color: #92400e; font-weight: bold; padding: 4px 8px; border-radius: 4px;">{value}</span></td>'
                else:
                    html += f'<td style="border: 1px solid #ddd; padding: 8px;">{value}</td>'
            else:
                # Alignement pour les nombres
                if isinstance(value, (int, float)):
                    html += f'<td style="border: 1px solid #ddd; padding: 8px; text-align: center;">{value}</td>'
                else:
                    html += f'<td style="border: 1px solid #ddd; padding: 8px;">{value}</td>'
        html += '</tr>'
    html += '</tbody>'
    html += '</table>'
    html += '</div>'
    
    return html

# ==================== SESSION STATE ====================
if 'resultats_affiches' not in st.session_state:
    st.session_state.resultats_affiches = None
if 'df_resultats' not in st.session_state:
    st.session_state.df_resultats = None
if 'erreurs_list' not in st.session_state:
    st.session_state.erreurs_list = None

# ==================== INTERFACE ====================

# Upload
col1, col2 = st.columns(2)

with col1:
    st.markdown("""
    <div class="card">
        <div class="card-title"><span>📊</span> Fichier Reply</div>
        <div class="upload-area">
    """, unsafe_allow_html=True)
    reply_file = st.file_uploader("reply.xlsx", type=['xlsx', 'xls'], label_visibility="collapsed")
    st.markdown('</div></div>', unsafe_allow_html=True)

with col2:
    st.markdown("""
    <div class="card">
        <div class="card-title"><span>📦</span> Fichiers Stock</div>
        <div class="upload-area">
    """, unsafe_allow_html=True)
    stock_files = st.file_uploader("Fichiers stock", type=['xlsx', 'xls'], accept_multiple_files=True, label_visibility="collapsed")
    st.markdown('</div></div>', unsafe_allow_html=True)

if reply_file and stock_files:
    
    with st.spinner("📥 Chargement..."):
        dict_reply = charger_feuilles_reply(reply_file)
        dict_stocks = charger_stocks(stock_files)
    
    if dict_reply and dict_stocks:
        
        st.markdown(f"""
        <div class="info-box">
            📁 <strong>{len(dict_reply)} feuille(s) trouvée(s) : {', '.join(dict_reply.keys())}</strong>
        </div>
        """, unsafe_allow_html=True)
        
        # IDL par modèle
        st.markdown("""
        <div class="card">
            <div class="card-title"><span>🔑</span> IDL par modèle</div>
        """, unsafe_allow_html=True)
        
        idl_par_modele = {}
        cols = st.columns(min(3, len(dict_reply)))
        
        for i, modele in enumerate(dict_reply.keys()):
            with cols[i % len(cols)]:
                st.markdown(f"**📱 {modele}**")
                idl = st.text_input("", key=f"idl_{modele}", placeholder="Entrez l'IDL", label_visibility="collapsed")
                if idl:
                    idl_par_modele[modele] = idl
        
        st.markdown('</div>', unsafe_allow_html=True)
        
        # Bouton centré
        col_btn1, col_btn2, col_btn3 = st.columns([1, 2, 1])
        with col_btn2:
            verifier = st.button("🚀 LANCER LA VÉRIFICATION", use_container_width=True)
        
        if verifier:
            if not idl_par_modele:
                st.warning("⚠️ Veuillez saisir au moins un IDL")
            else:
                resultats = []
                erreurs = []
                
                with st.spinner("⏳ Vérification en cours..."):
                    for modele, df_feuille in dict_reply.items():
                        if modele not in idl_par_modele:
                            continue
                        
                        df_std = extraire_colonnes_reply(df_feuille)
                        if df_std is None:
                            continue
                        
                        df_filtre = df_std[df_std['Remarks'].isin(['Missing', 'shortage'])]
                        
                        if df_filtre.empty:
                            continue
                        
                        idl = idl_par_modele[modele]
                        
                        for _, row in df_filtre.iterrows():
                            part_n = row['Part_N']
                            desc = row['Description'][:50] if pd.notna(row['Description']) else ""
                            qty_for = row['Qty_for']
                            packing_qty = row['Packing_qty']
                            oversent_frs = row['Oversent_FRS']
                            moka_file = row['Moka_file']
                            remarks = row['Remarks']
                            
                            stock_file = None
                            for fname in dict_stocks.keys():
                                if moka_file in fname.replace('.xlsx', '').replace('.xls', ''):
                                    stock_file = fname
                                    break
                            
                            if not stock_file:
                                erreurs.append(f"{part_n}: Fichier {moka_file} non trouvé")
                                resultats.append({
                                    'Modèle': modele, 
                                    'Part N': part_n, 
                                    'Description': desc,
                                    'Remarks': remarks, 
                                    'IDL': idl,
                                    'Qty for': qty_for, 
                                    'Packing Qty': packing_qty,
                                    'Oversent Stock': '-',
                                    'Oversent FRS': oversent_frs,
                                    'Oversent Calculé': '-',
                                    'Écart': '-',
                                    'Status': '❌'
                                })
                                continue
                            
                            try:
                                oversent_stock = get_oversent_stock(dict_stocks[stock_file], part_n, idl)
                                oversent_calc = oversent_stock + packing_qty - qty_for
                                ecart = oversent_calc - oversent_frs
                                correct = abs(ecart) < 0.01
                                
                                resultats.append({
                                    'Modèle': modele, 
                                    'Part N': part_n, 
                                    'Description': desc,
                                    'Remarks': remarks, 
                                    'IDL': idl,
                                    'Qty for': qty_for, 
                                    'Packing Qty': packing_qty,
                                    'Oversent Stock': oversent_stock,
                                    'Oversent FRS': oversent_frs,
                                    'Oversent Calculé': round(oversent_calc, 1),
                                    'Écart': round(ecart, 1),
                                    'Status': '✅' if correct else '❌'
                                })
                                
                            except Exception as e:
                                erreurs.append(f"{part_n}: {str(e)}")
                                resultats.append({
                                    'Modèle': modele, 
                                    'Part N': part_n, 
                                    'Description': desc,
                                    'Remarks': remarks, 
                                    'IDL': idl,
                                    'Qty for': qty_for, 
                                    'Packing Qty': packing_qty,
                                    'Oversent Stock': 'Erreur',
                                    'Oversent FRS': oversent_frs,
                                    'Oversent Calculé': 'Erreur',
                                    'Écart': 'Erreur',
                                    'Status': '⚠️'
                                })
                
                # Stocker les résultats
                st.session_state.resultats_affiches = True
                st.session_state.df_resultats = pd.DataFrame(resultats)
                st.session_state.erreurs_list = erreurs
        
        # Afficher les résultats
        if st.session_state.resultats_affiches and st.session_state.df_resultats is not None:
            df_res = st.session_state.df_resultats
            erreurs = st.session_state.erreurs_list
            
            st.markdown("---")
            st.markdown("""
            <div class="card">
                <div class="card-title"><span>📊</span> Résultats de la vérification</div>
            """, unsafe_allow_html=True)
            
            total = len(df_res)
            corrects = len(df_res[df_res['Status'] == '✅'])
            incorrects = len(df_res[df_res['Status'] == '❌'])
            taux = f"{corrects/total*100:.0f}" if total > 0 else "0"
            
            # Métriques
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.markdown(f'<div class="metric-total"><div class="metric-number">{total}</div><div class="metric-label">📊 TOTAL</div></div>', unsafe_allow_html=True)
            with col2:
                st.markdown(f'<div class="metric-correct"><div class="metric-number">{corrects}</div><div class="metric-label">✅ CORRECTS</div></div>', unsafe_allow_html=True)
            with col3:
                st.markdown(f'<div class="metric-incorrect"><div class="metric-number">{incorrects}</div><div class="metric-label">❌ INCORRECTS</div></div>', unsafe_allow_html=True)
            with col4:
                st.markdown(f'<div class="metric-taux"><div class="metric-number">{taux}%</div><div class="metric-label">🎯 TAUX</div></div>', unsafe_allow_html=True)
            
            st.markdown("<br>", unsafe_allow_html=True)
            
            # Afficher le tableau avec les nouveaux noms de colonnes
            html_table = afficher_tableau_html(df_res)
            st.markdown(html_table, unsafe_allow_html=True)
            
            # Export Excel (avec les nouveaux noms)
            df_export = df_res.copy()
            colonnes_rename_export = {
                'Qty for': 'Qty need in this lot',
                'Packing Qty': 'Qty sent in this lot',
                'Oversent Stock': 'Oversent in the previous IDL',
                'Oversent FRS': 'Oversent of reply',
                'Oversent Calculé': 'Oversent calculated',
                'Écart': 'GAP'
            }
            df_export = df_export.rename(columns=colonnes_rename_export)
            
            excel_data = exporter_excel_stylise(df_export, erreurs)
            
            col_btn1, col_btn2, col_btn3 = st.columns([1, 2, 1])
            with col_btn2:
                st.download_button(
                    label="📥 Télécharger Excel (avec bordures et couleurs)",
                    data=excel_data,
                    file_name="verification_resultats.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            
            if incorrects == 0:
                st.balloons()
                st.success("🎉 Félicitations ! Toutes les vérifications sont correctes !")
            
            st.markdown('</div>', unsafe_allow_html=True)

else:
    st.info("👈 Veuillez charger les fichiers pour commencer")

# Footer
st.markdown("""
<div class="footer">
    📐 Formule : Oversent réel = Oversent stock (ligne N-1) + Packing list qty - Qty for
</div>
""", unsafe_allow_html=True)
