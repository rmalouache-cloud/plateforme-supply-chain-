import streamlit as st
import pandas as pd
from io import BytesIO
from PIL import Image
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
import tempfile
import os
import time
import random

# =========================
# CONFIGURATION DE LA PAGE
# =========================
st.set_page_config(
    page_title="Packing Dashboard",
    page_icon="📦",
    layout="wide"
)

# =========================
# STYLES CSS PERSONNALISÉS
# =========================
st.markdown("""
    <style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 1rem;
    }
    .metric-card {
        background-color: #f8f9fa;
        padding: 1rem;
        border-radius: 10px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        text-align: center;
    }
    .info-text {
        background-color: #e3f2fd;
        padding: 1rem;
        border-radius: 10px;
        border-left: 5px solid #2196f3;
    }
    .error-text {
        background-color: #ffebee;
        padding: 1rem;
        border-radius: 10px;
        border-left: 5px solid #f44336;
        color: #c62828;
    }
    .download-spacing {
        margin-top: 50px;
    }
    
    /* Animation des ballons */
    @keyframes float {
        0% {
            transform: translateY(0) rotate(0deg);
            opacity: 1;
        }
        100% {
            transform: translateY(-100vh) rotate(360deg);
            opacity: 0;
        }
    }
    
    .balloon {
        position: fixed;
        bottom: -50px;
        font-size: 30px;
        pointer-events: none;
        z-index: 9999;
        animation: float 3s ease-in forwards;
    }
    
    @keyframes confetti {
        0% {
            transform: translateY(0) rotate(0deg);
            opacity: 1;
        }
        100% {
            transform: translateY(-100vh) rotate(720deg);
            opacity: 0;
        }
    }
    
    .confetti {
        position: fixed;
        bottom: -10px;
        font-size: 20px;
        pointer-events: none;
        z-index: 9998;
        animation: confetti 2.5s ease-out forwards;
    }
    </style>
""", unsafe_allow_html=True)

# Palette de couleurs pour les modèles
COLOR_PALETTE = [
    '#FFE5B4', '#FFD1DC', '#C4E0FA', '#D4F1F9', '#E6E6FA',
    '#C8E6C9', '#FFCCBC', '#F8BBD9', '#BBDEFB', '#C5E1A5',
    '#FFE0B2', '#D1C4E9', '#B2DFDB', '#F0F4C3', '#FFCDD2'
]

# Liste des émojis de ballons
BALLOON_EMOJIS = ['🎈', '🎈', '🎈', '🎈', '🎈', '🎈', '🎈', '🎈', '🎈', '🎈']
CONFETTI_EMOJIS = ['✨', '⭐', '🌟', '💫', '🎊', '🎉', '⚡', '💎', '🌸', '⭐']

def show_balloons_animation():
    """Affiche une animation de ballons colorés qui volent"""
    
    # Générer des ballons aléatoires
    balloon_html = ""
    for i in range(30):
        left_pos = random.randint(0, 100)
        delay = random.uniform(0, 1)
        size = random.randint(25, 45)
        emoji = random.choice(BALLOON_EMOJIS)
        rotation = random.randint(-30, 30)
        
        balloon_html += f"""
        <div class="balloon" style="
            left: {left_pos}%;
            animation-delay: {delay}s;
            font-size: {size}px;
            transform: rotate({rotation}deg);
        ">{emoji}</div>
        """
    
    # Ajouter des confettis
    for i in range(50):
        left_pos = random.randint(0, 100)
        delay = random.uniform(0, 1.5)
        size = random.randint(15, 25)
        emoji = random.choice(CONFETTI_EMOJIS)
        
        balloon_html += f"""
        <div class="confetti" style="
            left: {left_pos}%;
            animation-delay: {delay}s;
            font-size: {size}px;
        ">{emoji}</div>
        """
    
    # Afficher l'animation
    st.markdown(balloon_html, unsafe_allow_html=True)
    
    # Jouer l'animation pendant 3 secondes
    time.sleep(3)

def get_model_color(model_name, model_list):
    """Attribue une couleur unique à chaque modèle"""
    try:
        model_name_str = str(model_name)
        model_list_str = [str(m) for m in model_list]
        
        if model_name_str in model_list_str:
            index = model_list_str.index(model_name_str) % len(COLOR_PALETTE)
            return COLOR_PALETTE[index]
        return COLOR_PALETTE[0]
    except Exception:
        return COLOR_PALETTE[0]

def style_excel_with_borders_and_colors(file_path, model_list):
    """Applique des bordures et des couleurs par modèle au fichier Excel"""
    from openpyxl import load_workbook
    
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Définir les bordures
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Style pour l'en-tête
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F77B4", end_color="1F77B4", fill_type="solid")
    
    # Appliquer les bordures à toutes les cellules
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Style pour l'en-tête
            if cell.row == 1:
                cell.font = header_font
                cell.fill = header_fill
    
    # Appliquer les couleurs par modèle
    model_color_map = {}
    for idx, model in enumerate(model_list):
        model_color_map[str(model)] = get_model_color(model, model_list)
    
    # Colorier les lignes selon le modèle (colonne A)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        if row and row[0].value:
            model_value = str(row[0].value)
            if model_value in model_color_map:
                color_hex = model_color_map[model_value]
                color_rgb = color_hex.lstrip('#')
                model_fill = PatternFill(start_color=color_rgb, end_color=color_rgb, fill_type="solid")
                for cell in row:
                    cell.fill = model_fill
    
    wb.save(file_path)

# =========================
# CHARGEMENT DES LOGOS
# =========================
try:
    container_logo = Image.open("conteneur_logo.png")
    stream_logo = Image.open("stream_logo.png")
    logo_container_exists = True
except FileNotFoundError:
    logo_container_exists = False
    st.warning("⚠️ Les fichiers de logos sont introuvables. Vérifiez leur emplacement.")

# =========================
# EN-TÊTE AVEC LOGOS
# =========================
if logo_container_exists:
    col1, col2, col3 = st.columns([1.5, 3, 1.5])
    
    with col1:
        st.image(container_logo, width=200)
    
    with col2:
        st.markdown('<div class="main-header"> Packing Summary by Model & Type</div>', 
                   unsafe_allow_html=True)
    
    with col3:
        st.image(stream_logo, width=200)
else:
    st.markdown('<div class="main-header">📊 Packing Summary by Model & Type</div>', 
               unsafe_allow_html=True)

st.markdown("---")

# =========================
# SECTION INFORMATIONS D'EXPÉDITION
# =========================
with st.expander("📋 Informations d'expédition", expanded=True):
    col1, col2 = st.columns(2)
    
    with col1:
        apna = st.text_input(
            "🔖 Numéro APNA",
            placeholder="Ex: APNA-2024-001",
            help="Identifiant unique de l'expédition"
        )
    
    with col2:
        order_shipment = st.text_input(
            "📦 Order of Shipment",
            placeholder="Ex: ORD-001",
            help="Numéro d'ordre d'expédition"
        )

st.markdown("---")

# =========================
# TÉLÉCHARGEMENT DU FICHIER
# =========================
uploaded_file = st.file_uploader(
    "📥 Téléchargez votre fichier Excel",
    type=["xlsx"],
    help="Format accepté: .xlsx"
)

def find_column(df, possible_names):
    """Trouve une colonne dans le DataFrame en essayant plusieurs noms possibles"""
    for name in possible_names:
        for col in df.columns:
            if name.lower() in col.lower():
                return col
    return None

# Initialiser l'état de téléchargement dans session state
if 'download_clicked' not in st.session_state:
    st.session_state.download_clicked = False

if uploaded_file is not None:
    
    # =========================
    # CHARGEMENT ET NETTOYAGE
    # =========================
    with st.spinner("🔄 Chargement du fichier en cours..."):
        df = pd.read_excel(uploaded_file)
        
        # Nettoyage des noms de colonnes
        df.columns = (
            df.columns
            .astype(str)
            .str.strip()
            .str.replace("\n", " ", regex=False)
            .str.replace(r"\s+", " ", regex=True)
        )
        
        st.success("✅ Fichier chargé avec succès!")
    
    # =========================
    # DÉTECTION DES COLONNES REQUISES
    # =========================
    # Recherche de la colonne Model
    col_model = find_column(df, ['model', 'modele', 'modèle', 'product', 'article'])
    
    if col_model is None:
        st.markdown("""
        <div class="error-text">
            <strong>❌ Colonne 'Model' introuvable</strong><br>
            Votre fichier doit contenir une colonne pour les modèles (ex: 'Model', 'Modèle', 'Product').
        </div>
        """, unsafe_allow_html=True)
        st.stop()
    
    # Recherche de la colonne TYPE
    col_type = find_column(df, ['type', 'categorie', 'category'])
    
    if col_type is None:
        st.markdown("""
        <div class="error-text">
            <strong>❌ Colonne 'TYPE' introuvable</strong><br>
            Votre fichier doit contenir une colonne pour les types (ex: 'Type', 'Catégorie').
        </div>
        """, unsafe_allow_html=True)
        st.stop()
    
    # Recherche des colonnes numériques
    col_ctn = find_column(df, ['ctn', 'carton', 'box', 'quantity'])
    col_nw = find_column(df, ['n w', 'net', 'poids net', 'nw', 'weight net'])
    col_gw = find_column(df, ['g w', 'gross', 'poids brut', 'gw', 'weight gross'])
    col_vol = find_column(df, ['volume', 'cbm', 'vol', 'm3'])
    
    missing_cols = []
    if col_ctn is None: missing_cols.append("CTN/Quantité")
    if col_nw is None: missing_cols.append("N W/Poids Net")
    if col_gw is None: missing_cols.append("G W/Poids Brut")
    if col_vol is None: missing_cols.append("Volume/CBM")
    
    if missing_cols:
        st.markdown(f"""
        <div class="error-text">
            <strong>❌ Colonnes manquantes :</strong><br>
            {', '.join(missing_cols)}
        </div>
        """, unsafe_allow_html=True)
        st.stop()
    
    # Conversion en numérique
    for col in [col_ctn, col_nw, col_gw, col_vol]:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    
    # =========================
    # FILTRE PAR MODÈLE
    # =========================
    with st.sidebar:
        st.markdown("## 🔧 Filtres")
        models_all = df[col_model].dropna().unique()
        
        if len(models_all) == 0:
            st.error("Aucun modèle trouvé dans les données")
            st.stop()
        
        models_selected = st.multiselect(
            "🎯 Filtrer par modèle",
            options=sorted(models_all),
            default=sorted(models_all),
            help="Sélectionnez un ou plusieurs modèles"
        )
    
    df_filtered = df[df[col_model].isin(models_selected)]
    
    # =========================
    # SAISIE DES QUANTITÉS LOT
    # =========================
    st.subheader("📥 Saisie des quantités LOT par modèle")
    
    models = df_filtered[col_model].unique()
    lot_qty_dict = {}
    
    if len(models) > 0:
        # Création des colonnes dynamiquement
        cols = st.columns(min(len(models), 4))
        
        for idx, model in enumerate(models):
            col_idx = idx % 4
            lot_qty_dict[model] = cols[col_idx].number_input(
                f"🏷️ {model}",
                min_value=0,
                value=0,
                step=100,
                key=f"lot_{model}"
            )
    else:
        st.warning("Aucun modèle disponible après filtrage")
    
    # =========================
    # CALCUL DES AGRÉGATIONS
    # =========================
    with st.spinner("📊 Calcul des statistiques..."):
        result = df_filtered.groupby(
            [col_model, col_type], as_index=False
        ).agg({
            col_ctn: "sum",
            col_nw: "sum",
            col_gw: "sum",
            col_vol: "sum"
        })
        
        result.columns = [
            "Model", "TYPE",
            "CTN QTY",
            "TOTAL N W (KG)",
            "TOTAL G W (KG)",
            "TOTAL VOLUME (CBM)"
        ]
        
        result["LOT QTY"] = result["Model"].map(lot_qty_dict)
    
    # =========================
    # AFFICHAGE DU TABLEAU STYLISÉ AVEC COULEURS
    # =========================
    st.subheader("📦 Résumé détaillé")
    
    def style_table_with_colors(df):
        """Applique un style professionnel au DataFrame avec couleurs par modèle"""
        models_list = df['Model'].unique().tolist()
        
        # Fonction pour colorer les lignes selon le modèle
        def color_rows(row):
            model = row['Model']
            color = get_model_color(model, models_list)
            return [f'background-color: {color}' for _ in row]
        
        # Appliquer les couleurs et le style
        styled = df.style.apply(color_rows, axis=1)
        styled = styled.set_properties(**{
            'border': '1px solid #ddd',
            'text-align': 'center',
            'padding': '8px',
            'font-size': '14px'
        }).set_table_styles([
            {'selector': 'thead tr th',
             'props': [
                 ('background-color', '#1f77b4'),
                 ('color', 'white'),
                 ('border', '1px solid #1f77b4'),
                 ('text-align', 'center'),
                 ('padding', '10px'),
                 ('font-weight', 'bold')
             ]},
            {'selector': 'tbody tr:hover',
             'props': [('background-color', '#f5f5f5')]}
        ])
        
        # Formater les nombres
        return styled.format({
            'CTN QTY': '{:,.0f}',
            'TOTAL N W (KG)': '{:,.2f}',
            'TOTAL G W (KG)': '{:,.2f}',
            'TOTAL VOLUME (CBM)': '{:,.3f}',
            'LOT QTY': '{:,.0f}'
        })
    
    # Afficher le tableau avec couleurs
    st.dataframe(style_table_with_colors(result), use_container_width=True, height=400)
    
    # =========================
    # INDICATEURS CLÉS DE PERFORMANCE
    # =========================
    st.markdown("---")
    st.subheader("📈 Indicateurs globaux")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.markdown(f"""
        <div class="metric-card">
            <h3>📦</h3>
            <h4>Total CTN</h4>
            <h2>{int(result['CTN QTY'].sum()):,}</h2>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        st.markdown(f"""
        <div class="metric-card">
            <h3>⚖️</h3>
            <h4>Poids Net</h4>
            <h2>{result['TOTAL N W (KG)'].sum():,.2f} kg</h2>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        st.markdown(f"""
        <div class="metric-card">
            <h3>⚖️</h3>
            <h4>Poids Brut</h4>
            <h2>{result['TOTAL G W (KG)'].sum():,.2f} kg</h2>
        </div>
        """, unsafe_allow_html=True)
    
    with col4:
        st.markdown(f"""
        <div class="metric-card">
            <h3>📐</h3>
            <h4>Volume Total</h4>
            <h2>{result['TOTAL VOLUME (CBM)'].sum():,.3f} m³</h2>
        </div>
        """, unsafe_allow_html=True)
    
    # =========================
    # PRÉPARATION DU TÉLÉCHARGEMENT (AVEC BORDURES ET COULEURS)
    # =========================
    safe_apna = apna.strip().replace(" ", "_") if apna else "NO_APNA"
    safe_order = order_shipment.strip().replace(" ", "_") if order_shipment else "NO_ORDER"
    file_name = f"packing_summary_{safe_apna}_{safe_order}.xlsx"
    
    # Créer un fichier temporaire
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmpfile:
        temp_path = tmpfile.name
    
    # Sauvegarder l'Excel
    with pd.ExcelWriter(temp_path, engine='openpyxl') as writer:
        result.to_excel(writer, index=False, sheet_name="Summary")
        
        # Ajout d'une feuille de métadonnées
        metadata = pd.DataFrame({
            "Information": ["APNA", "Order of Shipment", "Date de génération", "Colonnes utilisées"],
            "Valeur": [
                apna if apna else "N/A", 
                order_shipment if order_shipment else "N/A",
                pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
                f"Model: {col_model}, Type: {col_type}, CTN: {col_ctn}, NW: {col_nw}, GW: {col_gw}, Vol: {col_vol}"
            ]
        })
        metadata.to_excel(writer, index=False, sheet_name="Metadata")
    
    # Appliquer les bordures et couleurs au fichier Excel
    style_excel_with_borders_and_colors(temp_path, result['Model'].unique().tolist())
    
    # Lire le fichier modifié
    with open(temp_path, 'rb') as f:
        excel_data = f.read()
    
    # Nettoyer le fichier temporaire
    os.unlink(temp_path)
    
    # =========================
    # BOUTON DE TÉLÉCHARGEMENT (AVEC ANIMATION DE BALLONS)
    # =========================
    # Ajout d'espace pour décaler le bouton vers le bas
    st.markdown('<div class="download-spacing"></div>', unsafe_allow_html=True)
    
    col_download1, col_download2, col_download3 = st.columns([1, 2, 1])
    with col_download2:
        # Bouton de téléchargement avec callback
        if st.download_button(
            label="📥 Télécharger le rapport Excel",
            data=excel_data,
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
            key="download_button"
        ):
            st.session_state.download_clicked = True
    
    # Afficher l'animation si le bouton a été cliqué
    if st.session_state.download_clicked:
        show_balloons_animation()
        # Réinitialiser l'état après l'animation
        st.session_state.download_clicked = False
        st.rerun()
    
    # =========================
    # SECTION INFORMATIONS SUPPLÉMENTAIRES
    # =========================
    with st.expander("ℹ️ Informations sur le rapport"):
        st.markdown(f"""
        <div class="info-text">
            <strong>📋 Détails du traitement :</strong><br>
            - Modèles sélectionnés : {', '.join(map(str, models_selected))}<br>
            - Lignes traitées : {len(df_filtered)}<br>
            - Date de génération : {pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")}<br>
            - Fichier source : {uploaded_file.name}
        </div>
        """, unsafe_allow_html=True)

else:
    # Affichage d'un message d'accueil lorsqu'aucun fichier n'est chargé
    st.markdown("""
    <div class="info-text" style="text-align: center;">
        <h3>👋 Bienvenue sur le Packing Dashboard</h3>
        <p>Pour commencer, veuillez télécharger votre fichier Excel via le bouton ci-dessus.</p>
        <p>📋 <strong>Format attendu :</strong> Fichier .xlsx contenant des colonnes pour :</p>
        <ul style="text-align: left; display: inline-block;">
            <li>Modèle (ex: "Model", "Modèle", "Product")</li>
            <li>Type (ex: "TYPE", "Type", "Catégorie")</li>
            <li>Quantité CTN (ex: "CTN", "QTY", "Quantity")</li>
            <li>Poids Net (ex: "N W", "Net Weight")</li>
            <li>Poids Brut (ex: "G W", "Gross Weight")</li>
            <li>Volume (ex: "Volume", "CBM")</li>
        </ul>
    </div>
    """, unsafe_allow_html=True)
    
    # Affichage d'un exemple
    with st.expander("📖 Voir un exemple du format attendu"):
        example_data = {
            "Model": ["MODEL_A", "MODEL_A", "MODEL_B"],
            "TYPE": ["Type1", "Type2", "Type1"],
            "CTN": [10, 5, 8],
            "N W": [100.5, 50.2, 80.0],
            "G W": [120.0, 60.5, 95.5],
            "VOLUME": [0.5, 0.25, 0.4]
        }
        example_df = pd.DataFrame(example_data)
        st.dataframe(example_df, use_container_width=True)
