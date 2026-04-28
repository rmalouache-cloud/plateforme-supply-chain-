import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side
import io
from typing import Tuple, Optional

# Configuration de la page
st.set_page_config(
    page_title="BOM Comparator",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# Style CSS personnalisé
st.markdown("""
    <style>
    .main-header {
        text-align: center;
        padding: 1rem;
        background: linear-gradient(90deg, #1e3c72, #2a5298);
        border-radius: 10px;
        margin-bottom: 2rem;
    }
    .status-badge {
        padding: 0.2rem 0.5rem;
        border-radius: 20px;
        font-weight: bold;
        text-align: center;
    }
    </style>
""", unsafe_allow_html=True)

# En-tête principal
st.markdown('<div class="main-header"><h1 style="color:white;">📊 BOM Comparison Tool</h1></div>', unsafe_allow_html=True)

# Upload des fichiers
col1, col2 = st.columns(2)
with col1:
    old_file = st.file_uploader("📂 **Ancienne BOM**", type=["xlsx"], help="Téléchargez l'ancienne version de la BOM")
with col2:
    new_file = st.file_uploader("📂 **Nouvelle BOM**", type=["xlsx"], help="Téléchargez la nouvelle version de la BOM")

# ============================================================================
# FONCTIONS UTILITAIRES
# ============================================================================

def safe_join(items: list) -> str:
    """Convertit une liste en chaîne de caractères séparée par des virgules"""
    if not isinstance(items, list):
        return ""
    return ", ".join(str(item) for item in items if pd.notna(item))

def find_boundary_indices(df: pd.DataFrame, start_keyword: str, end_keyword: str) -> Tuple[Optional[int], Optional[int]]:
    """Trouve les indices de début et fin dans un DataFrame"""
    start_idx = None
    end_idx = None
    
    for idx, desc in enumerate(df['Description']):
        desc_upper = str(desc).upper()
        if start_keyword in desc_upper and start_idx is None:
            start_idx = idx
        if end_keyword in desc_upper and end_idx is None:
            end_idx = idx
            
    return start_idx, end_idx

def extract_components_by_type(df: pd.DataFrame, bom_type: str) -> pd.DataFrame:
    """
    Extrait les composants CKD ou SKD selon la ligne de démarcation
    CKD: de "ASS'Y - MAIN BOARD（CKD）" jusqu'à "BARCODE LABEL"
    SKD: tout le reste
    """
    start_keyword = "ASS'Y - MAIN BOARD（CKD）".upper()
    end_keyword = "BARCODE LABEL".upper()
    
    start_idx, end_idx = find_boundary_indices(df, start_keyword, end_keyword)
    
    if bom_type == "CKD":
        if start_idx is not None:
            return df.iloc[start_idx:end_idx+1].copy() if end_idx is not None else df.iloc[start_idx:].copy()
        return pd.DataFrame()
    
    else:  # SKD
        if start_idx is not None:
            before_ckd = df.iloc[:start_idx].copy()
            after_ckd = df.iloc[end_idx+1:].copy() if end_idx is not None else pd.DataFrame()
            return pd.concat([before_ckd, after_ckd], ignore_index=True)
        return df.copy()

def clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Nettoie et prépare le DataFrame pour la comparaison"""
    cols = ["PN", "Description", "bom_qty", "BOM text"]
    df = df[cols].copy()
    df.rename(columns={"BOM text": "Position"}, inplace=True)
    
    # Nettoyage des données
    df["PN"] = df["PN"].astype(str).str.strip().str.upper()
    df["Description"] = df["Description"].astype(str).str.strip().str.upper()
    df["Position"] = df["Position"].astype(str).str.strip().str.upper()
    
    # Conversion des quantités
    df["bom_qty"] = (
        df["bom_qty"]
        .astype(str)
        .str.replace(",", ".", regex=False)
    )
    df["bom_qty"] = pd.to_numeric(df["bom_qty"], errors="coerce").fillna(0)
    
    # Agrégation par PN
    df = df.groupby(["PN"], as_index=False).agg({
        "Description": "first",
        "bom_qty": "sum",
        "Position": list
    })
    
    return df

def determine_status(row: pd.Series) -> str:
    """Détermine le statut de comparaison pour une ligne"""
    if row["_merge"] == "left_only":
        return "❌ Manquant dans nouvelle BOM"
    
    if row["_merge"] == "right_only":
        return "➕ Nouveau dans nouvelle BOM"
    
    qty_old, qty_new = row["bom_qty_old"], row["bom_qty_new"]
    pos_old = set(row["Position_old"]) if isinstance(row["Position_old"], list) else set()
    pos_new = set(row["Position_new"]) if isinstance(row["Position_new"], list) else set()
    
    if qty_old != qty_new:
        return "⚠️ Quantité différente"
    
    if pos_old != pos_new:
        return "📍 Position différente"
    
    return "✅ Conforme"

def run_comparison(old_df: pd.DataFrame, new_df: pd.DataFrame, component_type: str = "GENERAL") -> pd.DataFrame:
    """Exécute la comparaison entre deux DataFrames"""
    
    old = clean_dataframe(old_df)
    new = clean_dataframe(new_df)
    
    # Fusion des données
    df = old.merge(
        new,
        on="PN",
        how="outer",
        suffixes=("_old", "_new"),
        indicator=True
    )
    
    # Application du statut
    df["Status"] = df.apply(determine_status, axis=1)
    
    # Construction du résultat
    result = []
    for _, row in df.iterrows():
        pos_old = row.get("Position_old", []) if isinstance(row.get("Position_old"), list) else []
        pos_new = row.get("Position_new", []) if isinstance(row.get("Position_new"), list) else []
        
        result.append({
            "Type": component_type,
            "PN (Ancien)": row["PN"] if row["_merge"] != "right_only" else "",
            "Description (Ancien)": row.get("Description_old", ""),
            "Qté (Ancien)": row.get("bom_qty_old", 0),
            "Position (Ancien)": safe_join(pos_old),
            "PN (Nouveau)": row["PN"] if row["_merge"] != "left_only" else "",
            "Description (Nouveau)": row.get("Description_new", ""),
            "Qté (Nouveau)": row.get("bom_qty_new", 0),
            "Position (Nouveau)": safe_join(pos_new),
            "Statut": row["Status"]
        })
    
    return pd.DataFrame(result)

def apply_excel_styling(wb, colors: dict):
    """Applique la coloration conditionnelle et les bordures au fichier Excel"""
    ws = wb.active
    
    # Définir le style de bordure noire
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Appliquer les bordures à toutes les cellules utilisées
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
    
    # Trouver la colonne Statut
    status_col = None
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == "Statut":
            status_col = col
            break
    
    # Appliquer les couleurs de fond selon le statut
    if status_col:
        for row in range(2, ws.max_row + 1):
            status = ws.cell(row=row, column=status_col).value
            for key, color in colors.items():
                if key in status:
                    fill = PatternFill(start_color=color, fill_type="solid")
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).fill = fill
    
    # Ajuster la largeur des colonnes automatiquement
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Limiter à 50 caractères max
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Style pour l'en-tête (gras et centré)
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)
        cell.alignment = cell.alignment.copy(horizontal='center', vertical='center')

def create_download_button(final_result: pd.DataFrame, filename: str):
    """Crée un bouton de téléchargement avec mise en forme Excel"""
    output = io.BytesIO()
    final_result.to_excel(output, index=False)
    output.seek(0)
    
    wb = load_workbook(output)
    
    colors = {
        "Conforme": "C6EFCE",      # Vert
        "Manquant": "FFC7CE",      # Rouge
        "Nouveau": "FFC7CE",       # Rouge
        "Quantité différente": "FFEB9C",  # Jaune
        "Position différente": "BDD7EE"   # Bleu
    }
    
    apply_excel_styling(wb, colors)
    
    final_file = io.BytesIO()
    wb.save(final_file)
    final_file.seek(0)
    
    st.download_button(
        label="📥 **Télécharger le rapport Excel**",
        data=final_file,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

# ============================================================================
# INTERFACE PRINCIPALE
# ============================================================================

st.markdown("---")
st.subheader("🔧 **Sélectionnez le type de comparaison**")

# Boutons de comparaison
col1, col2, col3 = st.columns(3)

with col1:
    ckd_button = st.button("📟 **CKD**", use_container_width=True, help="Comparer uniquement les composants CKD")
with col2:
    skd_button = st.button("🔌 **SKD**", use_container_width=True, help="Comparer uniquement les composants SKD")
with col3:
    both_button = st.button("🖥️ **CKD + SKD**", use_container_width=True, help="Comparer tous les composants")

# Traitement de la comparaison
if any([ckd_button, skd_button, both_button]):
    
    if old_file is None or new_file is None:
        st.error("❌ **Veuillez télécharger les deux fichiers BOM avant de continuer**")
        st.stop()
    
    # Lecture des fichiers
    with st.spinner("📖 Lecture des fichiers en cours..."):
        old = pd.read_excel(old_file)
        new = pd.read_excel(new_file)
        old.columns = old.columns.str.strip()
        new.columns = new.columns.str.strip()
    
    all_results = []
    comparison_type = ""
    
    # Comparaison CKD
    if ckd_button:
        comparison_type = "CKD"
        st.info("📟 **Comparaison des composants CKD uniquement**")
        
        with st.spinner("🔄 Extraction des composants CKD..."):
            old_filtered = extract_components_by_type(old, "CKD")
            new_filtered = extract_components_by_type(new, "CKD")
        
        if old_filtered.empty and new_filtered.empty:
            st.warning("⚠️ Aucun composant CKD trouvé dans les fichiers")
        else:
            st.success(f"✅ Composants CKD trouvés: **{len(old_filtered)}** dans ancienne BOM, **{len(new_filtered)}** dans nouvelle BOM")
            result = run_comparison(old_filtered, new_filtered, "CKD")
            all_results.append(result)
    
    # Comparaison SKD
    elif skd_button:
        comparison_type = "SKD"
        st.info("🔌 **Comparaison des composants SKD uniquement**")
        
        with st.spinner("🔄 Extraction des composants SKD..."):
            old_filtered = extract_components_by_type(old, "SKD")
            new_filtered = extract_components_by_type(new, "SKD")
        
        if old_filtered.empty and new_filtered.empty:
            st.warning("⚠️ Aucun composant SKD trouvé dans les fichiers")
        else:
            st.success(f"✅ Composants SKD trouvés: **{len(old_filtered)}** dans ancienne BOM, **{len(new_filtered)}** dans nouvelle BOM")
            result = run_comparison(old_filtered, new_filtered, "SKD")
            all_results.append(result)
    
    # Comparaison complète
    elif both_button:
        comparison_type = "COMPLET"
        st.info("🖥️ **Comparaison complète CKD + SKD**")
        
        with st.spinner("🔄 Extraction des composants..."):
            old_ckd = extract_components_by_type(old, "CKD")
            old_skd = extract_components_by_type(old, "SKD")
            old_full = pd.concat([old_ckd, old_skd], ignore_index=True)
            
            new_ckd = extract_components_by_type(new, "CKD")
            new_skd = extract_components_by_type(new, "SKD")
            new_full = pd.concat([new_ckd, new_skd], ignore_index=True)
        
        st.success(f"✅ **Résumé de la comparaison:**")
        st.markdown(f"""
        - Total composants: **{len(old_full)}** dans ancienne BOM, **{len(new_full)}** dans nouvelle BOM
        - CKD: **{len(old_ckd)}** (ancien) / **{len(new_ckd)}** (nouveau)
        - SKD: **{len(old_skd)}** (ancien) / **{len(new_skd)}** (nouveau)
        """)
        
        result = run_comparison(old_full, new_full, "COMPLET")
        all_results.append(result)
    
    # Affichage des résultats
    if all_results:
        final_result = pd.concat(all_results, ignore_index=True)
        
        # Conversion en string pour éviter les crashes
        for col in final_result.columns:
            final_result[col] = final_result[col].astype(str)
        
        st.markdown("---")
        st.subheader("📊 **Résultats de la comparaison**")
        
        # Statistiques rapides
        col_stats1, col_stats2, col_stats3, col_stats4 = st.columns(4)
        with col_stats1:
            st.metric("📊 Total lignes", len(final_result))
        with col_stats2:
            conform = len(final_result[final_result['Statut'] == '✅ Conforme'])
            st.metric("✅ Conformes", conform)
        with col_stats3:
            missing = len(final_result[final_result['Statut'].str.contains('Manquant|Nouveau', na=False)])
            st.metric("❌ Différences", missing)
        with col_stats4:
            qty_diff = len(final_result[final_result['Statut'] == '⚠️ Quantité différente'])
            st.metric("⚠️ Qté différente", qty_diff)
        
        # Affichage du tableau
        st.dataframe(final_result, use_container_width=True, height=400)
        
        # Bouton de téléchargement
        filename = f"BOM_comparison_{comparison_type}_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        create_download_button(final_result, filename)
        
    else:
        st.error("❌ **Aucun résultat à afficher**")
