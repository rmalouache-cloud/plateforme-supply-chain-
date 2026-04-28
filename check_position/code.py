import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.styles.colors import Color
from PIL import Image
import os

st.set_page_config(page_title="CKD Position Validator", layout="wide")

# ==================== TRANSLATIONS ====================
def get_text(lang, key):
    """Get translated text based on selected language"""
    texts = {
        # French translations
        'fr': {
            'title': "Vérificateur de Positions CKD",
            'subtitle': "Vérifie que le nombre de positions correspond à la quantité pour les composants CKD uniquement",
            'upload': "📂 Upload votre fichier BOM",
            'preview': "📋 Aperçu du fichier chargé",
            'missing_cols': "❌ Colonnes manquantes:",
            'available_cols': "Les colonnes disponibles sont:",
            'extracting': "Extraction des composants CKD...",
            'no_ckd': "⚠️ Aucun composant CKD trouvé dans ce fichier",
            'help_descriptions': "Voici les 20 premières descriptions du fichier pour vérifier:",
            'extracted': "✅ {count} composants CKD extraits",
            'verify_btn': "🔍 Vérifier les positions CKD",
            'verifying': "Vérification en cours...",
            'summary': "📊 Résumé de la vérification CKD",
            'total': "📊 Total",
            'conforming': "✅ Conformes",
            'errors': "❌ Erreurs",
            'to_fix': "À corriger",
            'missing': "⚠️ Manque",
            'extra': "⚠️ Trop",
            'show_problems': "⚠️ Afficher uniquement les lignes non conformes",
            'details': "🔍 Détail de la vérification CKD",
            'non_conforming': "⚠️ {count} ligne(s) non conforme(s) sur {total} total",
            'no_problems': "✅ Aucune ligne non conforme trouvée !",
            'all_good': "✨ Toutes les {count} lignes sont conformes ou non applicables",
            'download': "📥 Télécharger le rapport Excel",
            'error': "❌ Erreur:",
            'upload_prompt': "👆 Veuillez uploader un fichier Excel",
            'language': "🌐 Language / Langue",
            'french': "Français",
            'english': "English",
            'logo_not_found': "Logo non trouvé:",
            'empty': "✅ VIDE",
            'error_no_position': "❌ ERREUR - Aucune position",
            'conforming_label': "✅ CONFORME",
            'missing_label': "⚠️ MANQUE - {count} position(s) manquante(s)",
            'extra_label': "⚠️ TROP - {count} position(s) en excès",
            'no_need': "📌 NO NEED / NOT APPLICABLE",
            'excel_conforming': "✔ CONFORME",
            'excel_error': "● ERREUR - Aucune position",
            'excel_missing': "⚠ MANQUE - {count} position(s) manquante(s)",
            'excel_extra': "⚠ TROP - {count} position(s) en excès",
            'excel_no_need': "◉ NO NEED / NOT APPLICABLE",
            'excel_empty': "○ VIDE",
        },
        # English translations
        'en': {
            'title': "CKD Position Validator",
            'subtitle': "Verifies that the number of positions matches the quantity for CKD components only",
            'upload': "📂 Upload your BOM file",
            'preview': "📋 File preview",
            'missing_cols': "❌ Missing columns:",
            'available_cols': "Available columns are:",
            'extracting': "Extracting CKD components...",
            'no_ckd': "⚠️ No CKD components found in this file",
            'help_descriptions': "Here are the first 20 descriptions to check:",
            'extracted': "✅ {count} CKD components extracted",
            'verify_btn': "🔍 Verify CKD Positions",
            'verifying': "Verification in progress...",
            'summary': "📊 CKD Verification Summary",
            'total': "📊 Total",
            'conforming': "✅ Conforming",
            'errors': "❌ Errors",
            'to_fix': "To fix",
            'missing': "⚠️ Missing",
            'extra': "⚠️ Extra",
            'show_problems': "⚠️ Show only non-conforming lines",
            'details': "🔍 CKD Verification Details",
            'non_conforming': "⚠️ {count} non-conforming line(s) out of {total} total",
            'no_problems': "✅ No non-conforming lines found!",
            'all_good': "✨ All {count} lines are conforming or not applicable",
            'download': "📥 Download Excel Report",
            'error': "❌ Error:",
            'upload_prompt': "👆 Please upload an Excel file",
            'language': "🌐 Language / Langue",
            'french': "Français",
            'english': "English",
            'logo_not_found': "Logo not found:",
            'empty': "✅ EMPTY",
            'error_no_position': "❌ ERROR - No position",
            'conforming_label': "✅ CONFORMING",
            'missing_label': "⚠️ MISSING - {count} position(s) missing",
            'extra_label': "⚠️ EXTRA - {count} extra position(s)",
            'no_need': "📌 NO NEED / NOT APPLICABLE",
            'excel_conforming': "✔ CONFORMING",
            'excel_error': "● ERROR - No position",
            'excel_missing': "⚠ MISSING - {count} position(s) missing",
            'excel_extra': "⚠ EXTRA - {count} extra position(s)",
            'excel_no_need': "◉ NO NEED / NOT APPLICABLE",
            'excel_empty': "○ EMPTY",
        }
    }
    return texts[lang].get(key, key)

# Initialize language in session state
if 'language' not in st.session_state:
    st.session_state.language = 'en'  # Default to English

# Custom CSS for Streamlit
st.markdown("""
<style>
    .dataframe {
        font-size: 14px;
        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
    }
    .stMetric {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        border-radius: 10px;
        padding: 10px;
    }
</style>
""", unsafe_allow_html=True)

# Language selector in sidebar
st.sidebar.markdown("---")
lang_option = st.sidebar.selectbox(
    get_text(st.session_state.language, 'language'),
    options=['🇬🇧 English', '🇫🇷 Français'],
    index=0 if st.session_state.language == 'en' else 1
)

# Update language based on selection
if lang_option == '🇬🇧 English' and st.session_state.language != 'en':
    st.session_state.language = 'en'
    st.rerun()
elif lang_option == '🇫🇷 Français' and st.session_state.language != 'fr':
    st.session_state.language = 'fr'
    st.rerun()

# Get current language
lang = st.session_state.language
t = lambda key: get_text(lang, key)

# Logo and title
col_logo, col_title = st.columns([1, 5])
with col_logo:
    logo_path = "logo.png"
    if os.path.exists(logo_path):
        logo = Image.open(logo_path)
        st.image(logo, width=120)
    else:
        st.markdown("# 📺")
        st.caption(f"{t('logo_not_found')} {logo_path}")
with col_title:
    st.markdown(f"# {t('title')}")
st.markdown(t('subtitle'))

# Initialize session state
if 'show_problems' not in st.session_state:
    st.session_state.show_problems = False
if 'results_df' not in st.session_state:
    st.session_state.results_df = None
if 'verification_done' not in st.session_state:
    st.session_state.verification_done = False
if 'current_file_name' not in st.session_state:
    st.session_state.current_file_name = None

# File upload
old_file = st.file_uploader(t('upload'), type=["xlsx"])

# Check if a new file has been uploaded
if old_file is not None:
    if st.session_state.current_file_name != old_file.name:
        st.session_state.results_df = None
        st.session_state.verification_done = False
        st.session_state.show_problems = False
        st.session_state.current_file_name = old_file.name

def extract_ckd_components(df):
    """Extract only CKD components"""
    start_idx = None
    for idx, desc in enumerate(df['Description']):
        if 'ASS\'Y - MAIN BOARD（CKD）' in str(desc).upper() or 'ASSY - MAIN BOARD（CKD）' in str(desc).upper():
            start_idx = idx
            break
    
    end_idx = None
    for idx, desc in enumerate(df['Description']):
        if 'BARCODE LABEL' in str(desc).upper():
            end_idx = idx
            break
    
    if start_idx is not None and end_idx is not None:
        return df.iloc[start_idx:end_idx+1].copy()
    elif start_idx is not None:
        return df.iloc[start_idx:].copy()
    else:
        return pd.DataFrame()

def safe_join(x):
    if not isinstance(x, list):
        return str(x) if pd.notna(x) else ""
    return ", ".join(str(i) for i in x if pd.notna(i))

def extract_positions(bom_text):
    positions = []
    if bom_text and str(bom_text) != "nan":
        bom_text_str = str(bom_text)
        raw_positions = bom_text_str.split(',')
        for pos in raw_positions:
            pos = pos.strip()
            pos = pos.replace('[', '').replace(']', '').replace("'", "").replace('"', "").strip()
            if pos and pos != "nan":
                positions.append(pos)
    return positions

def is_non_component(description):
    non_components = [
        'ASS\'Y - MAIN BOARD（CKD）',
        'ASSY - MAIN BOARD（CKD）',
        'ASS\'Y - MAIN BOARD',
        'ASSY - MAIN BOARD',
        'ASS\'Y - MAIN BOARD（SMT）',
        'ASSY - MAIN BOARD（SMT）',
        'PCB',
        'THERMAL CONDUCTIVE',
        'BARCODE LABEL'
    ]
    desc_upper = str(description).upper()
    for non_comp in non_components:
        if non_comp.upper() in desc_upper:
            return True
    return False

def validate_ckd_positions(df, lang):
    results = []
    t = lambda key: get_text(lang, key)
    
    for idx, row in df.iterrows():
        pn = row.get("PN", "")
        description = row.get("Description", "")
        bom_text = row.get("BOM text", "")
        qty = row.get("bom_qty", 0)
        
        is_non_comp = is_non_component(description)
        
        try:
            if isinstance(qty, str):
                qty = qty.replace(",", ".")
            qty = float(qty) if qty else 0
        except:
            qty = 0
        
        positions = extract_positions(bom_text)
        nb_positions = len(positions)
        positions_str = safe_join(positions)
        
        # For Streamlit display (colored emojis)
        if is_non_comp:
            result_display = t('no_need')
            result_class = "no-need"
        elif nb_positions == 0 and qty == 0:
            result_display = t('empty')
            result_class = "empty"
        elif nb_positions == 0 and qty > 0:
            result_display = t('error_no_position')
            result_class = "error"
        elif nb_positions == qty:
            result_display = t('conforming_label')
            result_class = "conforming"
        elif nb_positions < qty:
            result_display = t('missing_label').format(count=int(qty - nb_positions))
            result_class = "missing"
        else:
            result_display = t('extra_label').format(count=int(nb_positions - qty))
            result_class = "extra"
        
        # For Excel export (colorable symbols)
        if is_non_comp:
            result_excel = t('excel_no_need')
        elif nb_positions == 0 and qty == 0:
            result_excel = t('excel_empty')
        elif nb_positions == 0 and qty > 0:
            result_excel = t('excel_error')
        elif nb_positions == qty:
            result_excel = t('excel_conforming')
        elif nb_positions < qty:
            result_excel = t('excel_missing').format(count=int(qty - nb_positions))
        else:
            result_excel = t('excel_extra').format(count=int(nb_positions - qty))
        
        results.append({
            "PN": pn,
            "Description": description,
            "QTY": int(qty) if qty == int(qty) else qty,
            "Position": positions_str if positions_str else "—",
            "QTY Calculated": nb_positions,
            "Result_Display": result_display,
            "Result_Excel": result_excel,
            "Result_Class": result_class
        })
    
    return pd.DataFrame(results)

def export_to_colored_excel(df, filename):
    """Export to Excel with colored background and colored symbols"""
    output = io.BytesIO()
    
    # Prepare data for export (use Result_Excel)
    export_df = df[["PN", "Description", "QTY", "Position", "QTY Calculated", "Result_Excel"]].copy()
    export_df = export_df.rename(columns={"Result_Excel": "Result"})
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        export_df.to_excel(writer, sheet_name='Verification_CKD', index=False)
        
        workbook = writer.book
        worksheet = writer.sheets['Verification_CKD']
        
        # Process each row
        for row_idx in range(2, worksheet.max_row + 1):
            result_cell = worksheet.cell(row=row_idx, column=6)
            result_text = str(result_cell.value)
            
            # Determine colors
            if "CONFORMING" in result_text or "CONFORME" in result_text:
                fill_color = "C6EFCE"
                symbol_color = "006100"
            elif "ERROR" in result_text or "ERREUR" in result_text:
                fill_color = "FFC7CE"
                symbol_color = "9C0006"
            elif "MISSING" in result_text or "MANQUE" in result_text or "EXTRA" in result_text or "TROP" in result_text:
                fill_color = "FFEB9C"
                symbol_color = "9C6500"
            elif "NO NEED" in result_text:
                fill_color = "D9E1F2"
                symbol_color = "1A3A5C"
            elif "EMPTY" in result_text or "VIDE" in result_text:
                fill_color = "E2EFDA"
                symbol_color = "006100"
            else:
                fill_color = "FFFFFF"
                symbol_color = "000000"
            
            # Apply fill
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_idx, column=col)
                cell.fill = fill
                
                if col == 6:
                    text = str(cell.value)
                    if len(text) > 0:
                        first_char = text[0]
                        if first_char in ['✔', '●', '⚠', '◉', '○']:
                            cell.font = Font(color=symbol_color, bold=True)
                        else:
                            cell.font = Font(color="000000")
                else:
                    cell.font = Font(color="000000")
                
                cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Borders
            thin_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            for col in range(1, worksheet.max_column + 1):
                worksheet.cell(row=row_idx, column=col).border = thin_border
        
        # Header style
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        
        for col in range(1, worksheet.max_column + 1):
            header_cell = worksheet.cell(row=1, column=col)
            header_cell.fill = header_fill
            header_cell.font = header_font
            header_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Column widths
        column_widths = {'A': 22, 'B': 40, 'C': 10, 'D': 35, 'E': 16, 'F': 50}
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width
        
        worksheet.freeze_panes = 'A2'
    
    output.seek(0)
    return output

def color_result_css(val):
    """CSS style for Streamlit table - with gradients"""
    if "CONFORME" in str(val) or "CONFORMING" in str(val):
        return 'background: linear-gradient(135deg, #92D050 0%, #7CB83D 100%); color: white; font-weight: bold; border-radius: 5px; padding: 5px;'
    elif "ERREUR" in str(val) or "ERROR" in str(val):
        return 'background: linear-gradient(135deg, #FF6B6B 0%, #E55555 100%); color: white; font-weight: bold; border-radius: 5px; padding: 5px;'
    elif "MANQUE" in str(val) or "MISSING" in str(val):
        return 'background: linear-gradient(135deg, #FFB347 0%, #FF9500 100%); color: white; font-weight: bold; border-radius: 5px; padding: 5px;'
    elif "TROP" in str(val) or "EXTRA" in str(val):
        return 'background: linear-gradient(135deg, #FFB347 0%, #FF9500 100%); color: white; font-weight: bold; border-radius: 5px; padding: 5px;'
    elif "NO NEED" in str(val):
        return 'background: linear-gradient(135deg, #5B9BD5 0%, #3A7BC8 100%); color: white; font-weight: bold; border-radius: 5px; padding: 5px;'
    elif "VIDE" in str(val) or "EMPTY" in str(val):
        return 'background: linear-gradient(135deg, #C5E0B4 0%, #A8D08D 100%); color: #2C3E50; font-weight: bold; border-radius: 5px; padding: 5px;'
    return ''

if old_file:
    try:
        df = pd.read_excel(old_file)
        df.columns = df.columns.str.strip()
        
        with st.expander(t('preview')):
            st.dataframe(df.head(10), use_container_width=True)
        
        required_cols = ["bom_qty", "BOM text"]
        missing_cols = [col for col in required_cols if col not in df.columns]
        
        if missing_cols:
            st.error(f"{t('missing_cols')} {', '.join(missing_cols)}")
            st.info(f"{t('available_cols')} {', '.join(df.columns)}")
        else:
            with st.spinner(t('extracting')):
                ckd_df = extract_ckd_components(df)
            
            if ckd_df.empty:
                st.warning(t('no_ckd'))
                st.info(t('help_descriptions'))
                st.dataframe(df[["Description"]].head(20), use_container_width=True)
            else:
                st.success(t('extracted').format(count=len(ckd_df)))
                
                if st.button(t('verify_btn'), use_container_width=True):
                    with st.spinner(t('verifying')):
                        st.session_state.results_df = validate_ckd_positions(ckd_df, lang)
                        st.session_state.verification_done = True
                        st.session_state.show_problems = False
                        st.rerun()
                
                # Display results if verification has been done
                if st.session_state.verification_done and st.session_state.results_df is not None:
                    results_df = st.session_state.results_df
                    
                    # Statistics with style
                    st.subheader(t('summary'))
                    
                    col1, col2, col3, col4, col5 = st.columns(5)
                    total = len(results_df)
                    conforming = len(results_df[results_df["Result_Display"].str.contains("CONFORME|CONFORMING", na=False)])
                    errors = len(results_df[results_df["Result_Display"].str.contains("ERREUR|ERROR", na=False)])
                    missing = len(results_df[results_df["Result_Display"].str.contains("MANQUE|MISSING", na=False)])
                    extra = len(results_df[results_df["Result_Display"].str.contains("TROP|EXTRA", na=False)])
                    
                    with col1:
                        st.metric(t('total'), total)
                    with col2:
                        st.metric(t('conforming'), conforming)
                    with col3:
                        st.metric(t('errors'), errors, delta=t('to_fix') if errors > 0 else None)
                    with col4:
                        st.metric(t('missing'), missing)
                    with col5:
                        st.metric(t('extra'), extra)
                    
                    st.markdown("---")
                    
                    # Checkbox for filtering
                    show_problems = st.checkbox(
                        t('show_problems'), 
                        value=st.session_state.show_problems,
                        key="filter_checkbox"
                    )
                    
                    st.session_state.show_problems = show_problems
                    
                    st.subheader(t('details'))
                    
                    display_cols = ["PN", "Description", "QTY", "Position", "QTY Calculated", "Result_Display"]
                    
                    if st.session_state.show_problems:
                        filtered_df = results_df[
                            (results_df["Result_Display"].str.contains("ERREUR|ERROR", na=False)) |
                            (results_df["Result_Display"].str.contains("MANQUE|MISSING", na=False)) |
                            (results_df["Result_Display"].str.contains("TROP|EXTRA", na=False))
                        ][display_cols].copy()
                        filtered_df = filtered_df.rename(columns={"Result_Display": "Result"})
                        
                        if len(filtered_df) > 0:
                            st.warning(t('non_conforming').format(count=len(filtered_df), total=len(results_df)))
                            styled_filtered = filtered_df.style.map(color_result_css, subset=['Result'])
                            st.dataframe(styled_filtered, use_container_width=True)
                        else:
                            st.success(t('no_problems'))
                            st.balloons()
                            st.info(t('all_good').format(count=len(results_df)))
                    else:
                        all_df = results_df[display_cols].copy()
                        all_df = all_df.rename(columns={"Result_Display": "Result"})
                        styled_df = all_df.style.map(color_result_css, subset=['Result'])
                        st.dataframe(styled_df, use_container_width=True)
                    
                    # Export Excel with colored symbols
                    colored_excel = export_to_colored_excel(results_df, "verification_positions_CKD.xlsx")
                    
                    col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 1])
                    with col_btn2:
                        st.download_button(
                            t('download'),
                            colored_excel,
                            f"verification_CKD_{old_file.name.replace('.xlsx', '')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        
    except Exception as e:
        st.error(f"{t('error')} {str(e)}")
else:
    st.info(t('upload_prompt'))
    st.session_state.results_df = None
    st.session_state.verification_done = False
    st.session_state.current_file_name = None
